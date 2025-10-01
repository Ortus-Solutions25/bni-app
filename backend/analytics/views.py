"""
Analytics ViewSet - RESTful API for Matrix data and Comparisons
"""
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from chapters.models import Chapter
from reports.models import MonthlyReport
from members.models import Member
from analytics.models import Referral
from bni.services.comparison_service import ComparisonService


class MatrixViewSet(viewsets.ViewSet):
    """
    ViewSet for Matrix operations.

    Provides endpoints for:
    - Referral matrix with totals
    - One-to-one matrix with totals
    - Combination matrix with summaries
    """
    permission_classes = [AllowAny]  # TODO: Add proper authentication

    @action(detail=False, methods=['get'], url_path='referral')
    def referral_matrix(self, request, chapter_id=None, report_id=None):
        """
        Return referral matrix for a specific monthly report.

        Returns pre-processed matrix data with calculated summary columns:
        - given: Total referrals given by each member
        - received: Total referrals received by each member
        - unique_given: Number of unique members given referrals to
        - unique_received: Number of unique members received referrals from
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

            # Return the pre-processed matrix data with calculated summaries
            result = monthly_report.referral_matrix_data

            # Add calculated summary columns if data exists
            if result and 'members' in result and 'matrix' in result:
                members = result['members']
                matrix = result['matrix']

                # Calculate totals for referral matrix
                totals = {
                    'given': {},
                    'received': {},
                    'unique_given': {},
                    'unique_received': {}
                }

                # Calculate row totals (given)
                for i, giver in enumerate(members):
                    row_total = 0
                    unique_count = 0
                    for j, value in enumerate(matrix[i] if i < len(matrix) else []):
                        if value and isinstance(value, (int, float)):
                            row_total += value
                            if value > 0:
                                unique_count += 1
                    totals['given'][giver] = row_total
                    totals['unique_given'][giver] = unique_count

                # Calculate column totals (received)
                for j, receiver in enumerate(members):
                    col_total = 0
                    unique_count = 0
                    for i in range(len(matrix)):
                        if j < len(matrix[i]):
                            value = matrix[i][j]
                            if value and isinstance(value, (int, float)):
                                col_total += value
                                if value > 0:
                                    unique_count += 1
                    totals['received'][receiver] = col_total
                    totals['unique_received'][receiver] = unique_count

                result['totals'] = totals

            return Response(result)

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
            return Response(
                {'error': 'Chapter or monthly report not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Matrix generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='one-to-one')
    def one_to_one_matrix(self, request, chapter_id=None, report_id=None):
        """
        Return one-to-one matrix for a specific monthly report.

        Returns pre-processed matrix data with calculated summary columns:
        - given: Total OTOs initiated by each member
        - received: Total OTOs received by each member
        - unique_given: Number of unique members had OTOs with (as initiator)
        - unique_received: Number of unique members had OTOs with (as receiver)
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

            # Return the pre-processed matrix data with calculated summaries
            result = monthly_report.oto_matrix_data

            # Add calculated summary columns if data exists
            if result and 'members' in result and 'matrix' in result:
                members = result['members']
                matrix = result['matrix']

                # Calculate totals for OTO matrix
                totals = {
                    'given': {},
                    'received': {},
                    'unique_given': {},
                    'unique_received': {}
                }

                # Calculate row totals (given)
                for i, giver in enumerate(members):
                    row_total = 0
                    unique_count = 0
                    for j, value in enumerate(matrix[i] if i < len(matrix) else []):
                        if value and isinstance(value, (int, float)):
                            row_total += value
                            if value > 0:
                                unique_count += 1
                    totals['given'][giver] = row_total
                    totals['unique_given'][giver] = unique_count

                # Calculate column totals (received)
                for j, receiver in enumerate(members):
                    col_total = 0
                    unique_count = 0
                    for i in range(len(matrix)):
                        if j < len(matrix[i]):
                            value = matrix[i][j]
                            if value and isinstance(value, (int, float)):
                                col_total += value
                                if value > 0:
                                    unique_count += 1
                    totals['received'][receiver] = col_total
                    totals['unique_received'][receiver] = unique_count

                result['totals'] = totals

            return Response(result)

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
            return Response(
                {'error': 'Chapter or monthly report not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Matrix generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='combination')
    def combination_matrix(self, request, chapter_id=None, report_id=None):
        """
        Return combination matrix for a specific monthly report.

        Returns matrix showing which members have both referrals and OTOs.
        Values use numeric mapping:
        - 0 (or empty): Neither
        - 1: OTO only
        - 2: Referral only
        - 3: Both

        Includes summary counts for each category per member.
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

            # Return the pre-processed matrix data with calculated summaries
            result = monthly_report.combination_matrix_data

            # Add calculated summary columns if data exists
            if result and 'members' in result and 'matrix' in result:
                members = result['members']
                matrix = result['matrix']

                # Calculate combination counts for each member using simple numeric mapping
                # - (empty/0) = Neither, 1 = OTO only, 2 = Referral only, 3 = Both
                summaries = {
                    'neither': {},
                    'oto_only': {},
                    'referral_only': {},
                    'both': {}
                }

                for i, member in enumerate(members):
                    neither_count = 0
                    oto_only_count = 0
                    referral_only_count = 0
                    both_count = 0

                    row_data = matrix[i] if i < len(matrix) else []
                    for j, value in enumerate(row_data):
                        if i != j:  # Don't count self-relationships
                            # Convert value to integer for comparison
                            if isinstance(value, str):
                                if value == '-' or value == '' or value == '0':
                                    int_value = 0
                                else:
                                    try:
                                        int_value = int(value)
                                    except (ValueError, TypeError):
                                        int_value = 0
                            elif isinstance(value, (int, float)):
                                int_value = int(value)
                            else:
                                int_value = 0

                            # Count based on simple numeric mapping
                            if int_value == 0:
                                neither_count += 1
                            elif int_value == 1:
                                oto_only_count += 1
                            elif int_value == 2:
                                referral_only_count += 1
                            elif int_value == 3:
                                both_count += 1
                            else:
                                # Any other value counts as neither
                                neither_count += 1

                    summaries['neither'][member] = neither_count
                    summaries['oto_only'][member] = oto_only_count
                    summaries['referral_only'][member] = referral_only_count
                    summaries['both'][member] = both_count

                result['summaries'] = summaries

            return Response(result)

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
            return Response(
                {'error': 'Chapter or monthly report not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Matrix generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ComparisonViewSet(viewsets.ViewSet):
    """
    ViewSet for comparing monthly reports.

    Provides endpoints for:
    - Referral matrix comparison
    - One-to-one matrix comparison
    - Combination matrix comparison
    - Comprehensive report comparison
    """
    permission_classes = [AllowAny]  # TODO: Add proper authentication

    @action(detail=False, methods=['get'], url_path='referral')
    def compare_referral(self, request, chapter_id=None, report_id=None, previous_report_id=None):
        """
        Compare referral matrices between two monthly reports.

        Returns comparison data showing changes between reports.
        """
        try:
            # Get both reports
            current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
            previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

            # Perform comparison
            comparison = ComparisonService.compare_referral_matrices(
                current_report.referral_matrix_data,
                previous_report.referral_matrix_data
            )

            return Response({
                'current_report': {
                    'id': current_report.id,
                    'month_year': current_report.month_year
                },
                'previous_report': {
                    'id': previous_report.id,
                    'month_year': previous_report.month_year
                },
                'comparison': comparison
            }, status=status.HTTP_200_OK)

        except MonthlyReport.DoesNotExist:
            return Response(
                {'error': 'One or both reports not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to compare referral matrices: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='one-to-one')
    def compare_oto(self, request, chapter_id=None, report_id=None, previous_report_id=None):
        """
        Compare one-to-one matrices between two monthly reports.

        Returns comparison data showing changes in OTO activities.
        """
        try:
            # Get both reports
            current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
            previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

            # Perform comparison
            comparison = ComparisonService.compare_oto_matrices(
                current_report.oto_matrix_data,
                previous_report.oto_matrix_data
            )

            return Response({
                'current_report': {
                    'id': current_report.id,
                    'month_year': current_report.month_year
                },
                'previous_report': {
                    'id': previous_report.id,
                    'month_year': previous_report.month_year
                },
                'comparison': comparison
            }, status=status.HTTP_200_OK)

        except MonthlyReport.DoesNotExist:
            return Response(
                {'error': 'One or both reports not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to compare one-to-one matrices: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='combination')
    def compare_combination(self, request, chapter_id=None, report_id=None, previous_report_id=None):
        """
        Compare combination matrices between two monthly reports.

        Returns comparison showing changes in combined referral/OTO patterns.
        """
        try:
            # Get both reports
            current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
            previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

            # Perform comparison
            comparison = ComparisonService.compare_combination_matrices(
                current_report.combination_matrix_data,
                previous_report.combination_matrix_data
            )

            return Response({
                'current_report': {
                    'id': current_report.id,
                    'month_year': current_report.month_year
                },
                'previous_report': {
                    'id': previous_report.id,
                    'month_year': previous_report.month_year
                },
                'comparison': comparison
            }, status=status.HTTP_200_OK)

        except MonthlyReport.DoesNotExist:
            return Response(
                {'error': 'One or both reports not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to compare combination matrices: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='comprehensive')
    def compare_comprehensive(self, request, chapter_id=None, report_id=None, previous_report_id=None):
        """
        Get comprehensive comparison between two monthly reports.

        Includes all matrices and insights across referrals, OTOs, and combinations.
        """
        try:
            # Get both reports
            current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
            previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

            # Perform comprehensive comparison
            comparison_data = ComparisonService.compare_monthly_reports(
                current_report,
                previous_report
            )

            return Response(comparison_data, status=status.HTTP_200_OK)

        except MonthlyReport.DoesNotExist:
            return Response(
                {'error': 'One or both reports not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to compare reports: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='download-excel')
    def download_comparison_excel(self, request, chapter_id=None, report_id=None, previous_report_id=None):
        """
        Download Excel comparison showing combination matrix with aggregate columns.

        Columns include:
        - Combination matrix (current report)
        - Neither, OTO only, Referral only, Both counts
        - Current Referrals, Previous Referrals, Change in Referrals
        - Previous Neither, Change in Neither
        """
        try:
            # Get both reports
            chapter = Chapter.objects.get(id=chapter_id)
            current_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)
            previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter=chapter)

            # Get combination matrices
            current_matrix_data = current_report.combination_matrix_data
            previous_matrix_data = previous_report.combination_matrix_data

            if not current_matrix_data or not previous_matrix_data:
                return Response(
                    {'error': 'Matrix data not available for one or both reports'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            current_matrix = current_matrix_data['matrix']
            member_names = current_matrix_data.get('member_names', current_matrix_data.get('members', []))
            previous_matrix = previous_matrix_data['matrix']

            # Get members for referral counting
            members = Member.objects.filter(chapter=chapter, is_active=True).order_by('normalized_name')
            member_dict = {m.normalized_name: m for m in members}

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Combination Matrix Comparison'

            # Styles
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            center_align = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Green for improvements, red for declines
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

            # Headers
            # Row 1: Member names + aggregate column headers
            ws.cell(1, 1, 'Giver \\ Receiver').font = header_font
            ws.cell(1, 1).fill = header_fill
            ws.cell(1, 1).alignment = center_align

            # Member name columns
            for i, name in enumerate(member_names, start=2):
                ws.cell(1, i, name).font = header_font
                ws.cell(1, i).fill = header_fill
                ws.cell(1, i).alignment = center_align

            # Aggregate columns
            agg_start_col = len(member_names) + 2
            aggregate_headers = [
                'Neither:', 'OTO only:', 'Referral only:', 'OTO and Referral:',
                'Current Referral:', 'Last Referral:', 'Change in Referrals:',
                'Last Neither:', 'Change in Neither:'
            ]
            for i, header in enumerate(aggregate_headers):
                col = agg_start_col + i
                ws.cell(1, col, header).font = header_font
                ws.cell(1, col).fill = header_fill
                ws.cell(1, col).alignment = center_align

            # Data rows
            for row_idx, giver_name in enumerate(member_names, start=2):
                # Giver name in column 1
                ws.cell(row_idx, 1, giver_name).font = Font(bold=True)
                ws.cell(row_idx, 1).alignment = center_align

                # Matrix values
                matrix_row = current_matrix[row_idx - 2]
                for col_idx, value in enumerate(matrix_row, start=2):
                    ws.cell(row_idx, col_idx, value).alignment = center_align

                # Calculate aggregates
                neither_count = sum(1 for v in matrix_row if v == 0)
                oto_only_count = sum(1 for v in matrix_row if v == 1)
                ref_only_count = sum(1 for v in matrix_row if v == 2)
                both_count = sum(1 for v in matrix_row if v == 3)

                # Current referrals (values 2 or 3 in current matrix)
                current_referrals = sum(1 for v in matrix_row if v in [2, 3])

                # Previous referrals and neither
                prev_matrix_row = previous_matrix[row_idx - 2]
                previous_referrals = sum(1 for v in prev_matrix_row if v in [2, 3])
                previous_neither = sum(1 for v in prev_matrix_row if v == 0)

                # Changes
                change_in_referrals = current_referrals - previous_referrals
                change_in_neither = neither_count - previous_neither

                # Add arrows to changes
                if change_in_referrals > 0:
                    ref_change_text = f"{change_in_referrals} ↗️"
                    ref_change_fill = green_fill
                elif change_in_referrals < 0:
                    ref_change_text = f"{change_in_referrals} ↘️"
                    ref_change_fill = red_fill
                else:
                    ref_change_text = f"{change_in_referrals} ➡️"
                    ref_change_fill = None

                if change_in_neither < 0:  # Decrease in neither is good
                    neither_change_text = f"{change_in_neither} ↘️"
                    neither_change_fill = green_fill
                elif change_in_neither > 0:
                    neither_change_text = f"{change_in_neither} ↗️"
                    neither_change_fill = red_fill
                else:
                    neither_change_text = f"{change_in_neither} ➡️"
                    neither_change_fill = None

                # Write aggregate data
                ws.cell(row_idx, agg_start_col, neither_count).alignment = center_align
                ws.cell(row_idx, agg_start_col + 1, oto_only_count).alignment = center_align
                ws.cell(row_idx, agg_start_col + 2, ref_only_count).alignment = center_align
                ws.cell(row_idx, agg_start_col + 3, both_count).alignment = center_align
                ws.cell(row_idx, agg_start_col + 4, current_referrals).alignment = center_align
                ws.cell(row_idx, agg_start_col + 5, previous_referrals).alignment = center_align

                # Change in referrals with color
                cell = ws.cell(row_idx, agg_start_col + 6, ref_change_text)
                cell.alignment = center_align
                if ref_change_fill:
                    cell.fill = ref_change_fill

                # Previous neither and change in neither
                ws.cell(row_idx, agg_start_col + 7, previous_neither).alignment = center_align
                cell = ws.cell(row_idx, agg_start_col + 8, neither_change_text)
                cell.alignment = center_align
                if neither_change_fill:
                    cell.fill = neither_change_fill

            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            for col_idx in range(2, len(member_names) + 2):
                ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 12
            for col_idx in range(agg_start_col, agg_start_col + len(aggregate_headers)):
                ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

            # Generate filename
            filename = f"{chapter.name.replace(' ', '_')}_comparison_{previous_report.month_year}_vs_{current_report.month_year}.xlsx"

            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # Save workbook to response
            wb.save(response)
            return response

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
            return Response(
                {'error': 'Chapter or reports not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Failed to generate Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
