from rest_framework import status, serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import models
from datetime import datetime, date
import tempfile
import os
from pathlib import Path

from chapters.models import Chapter
from members.models import Member
from reports.models import MonthlyReport, MemberMonthlyStats
from analytics.models import Referral, OneToOne, TYFCB, DataImportSession
from bni.serializers import ChapterSerializer, MemberSerializer, MemberCreateSerializer, MemberUpdateSerializer
from bni.services.excel_processor import ExcelProcessorService
from bni.services.matrix_generator import MatrixGenerator, DataValidator
from bni.services.bulk_upload_service import BulkUploadService
from bni.services.chapter_service import ChapterService
from bni.services.member_service import MemberService
from bni.services.comparison_service import ComparisonService


class FileUploadSerializer(serializers.Serializer):
    slip_audit_file = serializers.FileField()
    member_names_file = serializers.FileField(required=False)
    chapter_id = serializers.IntegerField()
    month_year = serializers.CharField(max_length=7, help_text="e.g., '2024-06'")
    upload_option = serializers.ChoiceField(choices=['slip_only', 'slip_and_members'], default='slip_only')


class ExcelUploadView(APIView):
    """Handle Excel file upload and processing."""

    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'error': 'Invalid data', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get validated data
            slip_audit_file = serializer.validated_data['slip_audit_file']
            member_names_file = serializer.validated_data.get('member_names_file')
            chapter_id = serializer.validated_data['chapter_id']
            month_year = serializer.validated_data['month_year']
            upload_option = serializer.validated_data['upload_option']

            # Validate chapter access
            try:
                chapter = Chapter.objects.get(id=chapter_id)
                # Add permission check here if needed
            except Chapter.DoesNotExist:
                return Response(
                    {'error': 'Chapter not found'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Validate file types
            if not slip_audit_file.name.lower().endswith(('.xls', '.xlsx')):
                return Response(
                    {'error': 'Only .xls and .xlsx files are supported for slip audit file'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if member_names_file and not member_names_file.name.lower().endswith(('.xls', '.xlsx')):
                return Response(
                    {'error': 'Only .xls and .xlsx files are supported for member names file'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process using the new monthly report method
            processor = ExcelProcessorService(chapter)
            result = processor.process_monthly_report(
                slip_audit_file=slip_audit_file,
                member_names_file=member_names_file,
                month_year=month_year
            )

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Processing failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BulkUploadView(APIView):
    """Handle Regional PALMS Summary bulk upload."""
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [AllowAny]

    def post(self, request):
        """Process Regional PALMS Summary report."""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']

        # Validate file type
        if not file.name.endswith(('.xls', '.xlsx')):
            return Response(
                {'error': 'Invalid file type. Please upload .xls or .xlsx file'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Process the file
            service = BulkUploadService()
            result = service.process_region_summary(file)

            if result['success']:
                return Response({
                    'success': True,
                    'message': 'Bulk upload completed successfully',
                    'details': {
                        'chapters_created': result['chapters_created'],
                        'chapters_updated': result['chapters_updated'],
                        'members_created': result['members_created'],
                        'members_updated': result['members_updated'],
                        'total_processed': result['total_processed'],
                        'warnings': result['warnings']
                    }
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'success': False,
                    'message': result.get('error', 'Processing failed'),
                    'details': {
                        'errors': result.get('errors', []),
                        'warnings': result.get('warnings', [])
                    }
                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            logger.exception("Error in bulk upload")
            return Response(
                {'error': f'Processing failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_referral_matrix(request, chapter_id, report_id):
    """Return referral matrix for a specific monthly report."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

        # Return the pre-processed matrix data with calculated summaries
        result = monthly_report.referral_matrix_data

        # Add calculated summary columns if data exists
        if result and 'members' in result and 'matrix' in result:
            members = result['members']
            matrix = result['matrix']

            # Calculate totals for referral matrix
            totals = {
                'given': {},
                'received': {},
                'unique_given': {},
                'unique_received': {}
            }

            # Calculate row totals (given)
            for i, giver in enumerate(members):
                row_total = 0
                unique_count = 0
                for j, value in enumerate(matrix[i] if i < len(matrix) else []):
                    if value and isinstance(value, (int, float)):
                        row_total += value
                        if value > 0:
                            unique_count += 1
                totals['given'][giver] = row_total
                totals['unique_given'][giver] = unique_count

            # Calculate column totals (received)
            for j, receiver in enumerate(members):
                col_total = 0
                unique_count = 0
                for i in range(len(matrix)):
                    if j < len(matrix[i]):
                        value = matrix[i][j]
                        if value and isinstance(value, (int, float)):
                            col_total += value
                            if value > 0:
                                unique_count += 1
                totals['received'][receiver] = col_total
                totals['unique_received'][receiver] = unique_count

            result['totals'] = totals

        return Response(result)

    except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
        return Response(
            {'error': 'Chapter or monthly report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Matrix generation failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_one_to_one_matrix(request, chapter_id, report_id):
    """Return one-to-one matrix for a specific monthly report."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

        # Return the pre-processed matrix data with calculated summaries
        result = monthly_report.oto_matrix_data

        # Add calculated summary columns if data exists
        if result and 'members' in result and 'matrix' in result:
            members = result['members']
            matrix = result['matrix']

            # Calculate totals for OTO matrix
            totals = {
                'given': {},
                'received': {},
                'unique_given': {},
                'unique_received': {}
            }

            # Calculate row totals (given)
            for i, giver in enumerate(members):
                row_total = 0
                unique_count = 0
                for j, value in enumerate(matrix[i] if i < len(matrix) else []):
                    if value and isinstance(value, (int, float)):
                        row_total += value
                        if value > 0:
                            unique_count += 1
                totals['given'][giver] = row_total
                totals['unique_given'][giver] = unique_count

            # Calculate column totals (received)
            for j, receiver in enumerate(members):
                col_total = 0
                unique_count = 0
                for i in range(len(matrix)):
                    if j < len(matrix[i]):
                        value = matrix[i][j]
                        if value and isinstance(value, (int, float)):
                            col_total += value
                            if value > 0:
                                unique_count += 1
                totals['received'][receiver] = col_total
                totals['unique_received'][receiver] = unique_count

            result['totals'] = totals

        return Response(result)

    except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
        return Response(
            {'error': 'Chapter or monthly report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Matrix generation failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_combination_matrix(request, chapter_id, report_id):
    """Return combination matrix for a specific monthly report."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

        # Return the pre-processed matrix data with calculated summaries
        result = monthly_report.combination_matrix_data

        # Add calculated summary columns if data exists
        if result and 'members' in result and 'matrix' in result:
            members = result['members']
            matrix = result['matrix']

            # Calculate combination counts for each member using simple numeric mapping
            # - (empty/0) = Neither, 1 = OTO only, 2 = Referral only, 3 = Both
            summaries = {
                'neither': {},
                'oto_only': {},
                'referral_only': {},
                'both': {}
            }

            for i, member in enumerate(members):
                neither_count = 0
                oto_only_count = 0
                referral_only_count = 0
                both_count = 0

                row_data = matrix[i] if i < len(matrix) else []
                for j, value in enumerate(row_data):
                    if i != j:  # Don't count self-relationships
                        # Convert value to integer for comparison
                        if isinstance(value, str):
                            if value == '-' or value == '' or value == '0':
                                int_value = 0
                            else:
                                try:
                                    int_value = int(value)
                                except (ValueError, TypeError):
                                    int_value = 0
                        elif isinstance(value, (int, float)):
                            int_value = int(value)
                        else:
                            int_value = 0

                        # Count based on simple numeric mapping
                        if int_value == 0:
                            neither_count += 1
                        elif int_value == 1:
                            oto_only_count += 1
                        elif int_value == 2:
                            referral_only_count += 1
                        elif int_value == 3:
                            both_count += 1
                        else:
                            # Any other value counts as neither
                            neither_count += 1

                summaries['neither'][member] = neither_count
                summaries['oto_only'][member] = oto_only_count
                summaries['referral_only'][member] = referral_only_count
                summaries['both'][member] = both_count

            result['summaries'] = summaries

        return Response(result)

    except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
        return Response(
            {'error': 'Chapter or monthly report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Matrix generation failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_member_summary(request, chapter_id):
    """Get comprehensive member summary for a chapter."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        members = list(Member.objects.filter(chapter=chapter, is_active=True))
        referrals = list(Referral.objects.filter(giver__chapter=chapter))
        one_to_ones = list(OneToOne.objects.filter(member1__chapter=chapter))
        tyfcbs = list(TYFCB.objects.filter(receiver__chapter=chapter))

        generator = MatrixGenerator(members)
        summary = generator.generate_member_summary(referrals, one_to_ones, tyfcbs)

        # Convert DataFrame to list of dictionaries
        result = summary.to_dict('records')

        return Response(result)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Summary generation failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_tyfcb_summary(request, chapter_id):
    """Get TYFCB summary for a chapter."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        members = list(Member.objects.filter(chapter=chapter, is_active=True))
        tyfcbs = list(TYFCB.objects.filter(receiver__chapter=chapter))

        generator = MatrixGenerator(members)
        summary = generator.generate_tyfcb_summary(tyfcbs)

        result = summary.to_dict('records')

        return Response(result)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'TYFCB summary failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_data_quality_report(request, chapter_id):
    """Get data quality report for a chapter."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        referrals = list(Referral.objects.filter(giver__chapter=chapter))
        one_to_ones = list(OneToOne.objects.filter(member1__chapter=chapter))
        tyfcbs = list(TYFCB.objects.filter(receiver__chapter=chapter))

        report = DataValidator.generate_quality_report(
            referrals, one_to_ones, tyfcbs
        )

        return Response(report)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Quality report failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_import_history(request, chapter_id):
    """Get import history for a chapter."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        sessions = DataImportSession.objects.filter(chapter=chapter)[:50]

        result = []
        for session in sessions:
            result.append({
                'id': session.id,
                'file_name': session.file_name,
                'import_date': session.import_date,
                'success': session.success,
                'records_processed': session.records_processed,
                'referrals_created': session.referrals_created,
                'one_to_ones_created': session.one_to_ones_created,
                'tyfcbs_created': session.tyfcbs_created,
                'errors_count': session.errors_count,
            })

        return Response(result)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Import history failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ========================================
# CHAPTER-FOCUSED API ENDPOINTS
# ========================================

@api_view(['GET'])
@permission_classes([AllowAny])
def chapter_dashboard(request):
    """Get dashboard data for all chapters."""
    try:
        from django.db.models import Count, Sum

        chapters = Chapter.objects.all().order_by('name')
        dashboard_data = []

        for chapter in chapters:
            # Get actual data from database
            members = Member.objects.filter(chapter=chapter, is_active=True)
            referrals = Referral.objects.filter(giver__chapter=chapter)
            tyfcbs = TYFCB.objects.filter(receiver__chapter=chapter)
            monthly_reports = MonthlyReport.objects.filter(chapter=chapter).order_by('-month_year')

            # Get latest data date from monthly report
            latest_date = None
            if monthly_reports.exists():
                latest_report = monthly_reports.first()
                latest_date = f"{latest_report.month_year}"

            # Calculate total TYFCB amount
            total_tyfcb = tyfcbs.aggregate(total=Sum('amount'))['total'] or 0.0

            # Calculate averages
            member_count = members.count()
            referral_count = referrals.count()
            one_to_ones = OneToOne.objects.filter(member1__chapter=chapter)
            oto_count = one_to_ones.count()

            avg_referrals = round(referral_count / member_count, 2) if member_count > 0 else 0
            avg_tyfcb = round(float(total_tyfcb) / member_count, 2) if member_count > 0 else 0
            avg_otos = round(oto_count / member_count, 2) if member_count > 0 else 0

            # Include member details for admin dashboard
            members_list = []
            for member in members:
                members_list.append({
                    'id': member.id,
                    'name': f"{member.first_name} {member.last_name}",
                    'first_name': member.first_name,
                    'last_name': member.last_name,
                    'business_name': member.business_name or '',
                    'classification': member.classification or '',
                    'is_active': member.is_active
                })

            chapter_data = {
                'id': chapter.id,
                'name': chapter.name,
                'location': chapter.location if hasattr(chapter, 'location') else 'Dubai',
                'latest_data_date': latest_date,
                'member_count': member_count,
                'members': members_list,
                'total_referrals': referral_count,
                'total_one_to_ones': oto_count,
                'total_tyfcb': float(total_tyfcb),
                'avg_referrals_per_member': avg_referrals,
                'avg_tyfcb_per_member': avg_tyfcb,
                'avg_otos_per_member': avg_otos,
                'has_data': monthly_reports.exists()
            }
            dashboard_data.append(chapter_data)

        return Response({
            'chapters': dashboard_data,
            'total_chapters': len(dashboard_data)
        })

    except Exception as e:
        return Response(
            {'error': f'Dashboard failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def chapter_detail(request, chapter_id):
    """Get detailed information for a specific chapter."""
    try:
        from django.db.models import Sum

        chapter = Chapter.objects.get(id=chapter_id)
        members = Member.objects.filter(chapter=chapter, is_active=True)

        # Get actual data from database
        referrals = Referral.objects.filter(giver__chapter=chapter)
        one_to_ones = OneToOne.objects.filter(member1__chapter=chapter)
        tyfcbs = TYFCB.objects.filter(receiver__chapter=chapter)
        monthly_reports = MonthlyReport.objects.filter(chapter=chapter).order_by('-month_year')

        # Get latest data date
        latest_date = None
        if monthly_reports.exists():
            latest_report = monthly_reports.first()
            latest_date = f"{latest_report.month_year}"

        # Calculate totals
        member_count = members.count()
        referral_count = referrals.count()
        oto_count = one_to_ones.count()
        total_tyfcb = tyfcbs.aggregate(total=Sum('amount'))['total'] or 0.0

        # Calculate averages
        avg_referrals = round(referral_count / member_count, 2) if member_count > 0 else 0
        avg_tyfcb = round(float(total_tyfcb) / member_count, 2) if member_count > 0 else 0
        avg_otos = round(oto_count / member_count, 2) if member_count > 0 else 0

        chapter_data = {
            'id': chapter.id,
            'name': chapter.name,
            'location': chapter.location if hasattr(chapter, 'location') else 'Dubai',
            'latest_data_date': latest_date,
            'member_count': member_count,
            'total_referrals': referral_count,
            'total_one_to_ones': oto_count,
            'total_tyfcb': float(total_tyfcb),
            'avg_referrals_per_member': avg_referrals,
            'avg_tyfcb_per_member': avg_tyfcb,
            'avg_otos_per_member': avg_otos,
            'has_data': monthly_reports.exists(),
            'members': [
                {
                    'id': member.id,
                    'full_name': member.full_name,
                    'business_name': member.business_name,
                    'classification': member.classification,
                    'email': member.email,
                    'phone': member.phone
                }
                for member in members
            ]
        }

        return Response(chapter_data)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Chapter detail failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_monthly_reports(request, chapter_id):
    """Get all monthly reports for a specific chapter."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_reports = MonthlyReport.objects.filter(chapter=chapter).order_by('-month_year')

        result = []
        for report in monthly_reports:
            result.append({
                'id': report.id,
                'month_year': report.month_year,
                'uploaded_at': report.uploaded_at,
                'processed_at': report.processed_at,
                'slip_audit_file': report.slip_audit_file.name if report.slip_audit_file else None,
                'member_names_file': report.member_names_file.name if report.member_names_file else None,
                'has_referral_matrix': bool(report.referral_matrix_data),
                'has_oto_matrix': bool(report.oto_matrix_data),
                'has_combination_matrix': bool(report.combination_matrix_data)
            })

        return Response(result)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Monthly reports retrieval failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_monthly_report(request, chapter_id, report_id):
    """Delete a monthly report."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

        # Delete associated files
        if monthly_report.slip_audit_file:
            monthly_report.slip_audit_file.delete(save=False)
        if monthly_report.member_names_file:
            monthly_report.member_names_file.delete(save=False)

        # Delete the report
        monthly_report.delete()

        return Response({'message': 'Monthly report deleted successfully'})

    except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
        return Response(
            {'error': 'Chapter or monthly report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Monthly report deletion failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_member_detail(request, chapter_id, report_id, member_id):
    """Get detailed member information including missing interaction lists."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)
        member = Member.objects.get(id=member_id, chapter=chapter)

        try:
            member_stats = MemberMonthlyStats.objects.get(
                member=member,
                monthly_report=monthly_report
            )
        except MemberMonthlyStats.DoesNotExist:
            # If no stats exist, return basic member info with empty lists
            member_stats = None

        # Get all chapter members for name resolution
        chapter_members = Member.objects.filter(chapter=chapter, is_active=True)
        member_lookup = {m.id: m.full_name for m in chapter_members}

        result = {
            'member': {
                'id': member.id,
                'full_name': member.full_name,
                'first_name': member.first_name,
                'last_name': member.last_name,
                'business_name': member.business_name,
                'classification': member.classification,
                'email': member.email,
                'phone': member.phone
            },
            'stats': {
                'referrals_given': member_stats.referrals_given if member_stats else 0,
                'referrals_received': member_stats.referrals_received if member_stats else 0,
                'one_to_ones_completed': member_stats.one_to_ones_completed if member_stats else 0,
                'tyfcb_inside_amount': float(member_stats.tyfcb_inside_amount) if member_stats else 0.0,
                'tyfcb_outside_amount': float(member_stats.tyfcb_outside_amount) if member_stats else 0.0
            },
            'missing_interactions': {
                'missing_otos': [
                    {
                        'id': member_id,
                        'name': member_lookup.get(member_id, 'Unknown')
                    }
                    for member_id in (member_stats.missing_otos if member_stats else [])
                    if member_id in member_lookup
                ],
                'missing_referrals_given_to': [
                    {
                        'id': member_id,
                        'name': member_lookup.get(member_id, 'Unknown')
                    }
                    for member_id in (member_stats.missing_referrals_given_to if member_stats else [])
                    if member_id in member_lookup
                ],
                'missing_referrals_received_from': [
                    {
                        'id': member_id,
                        'name': member_lookup.get(member_id, 'Unknown')
                    }
                    for member_id in (member_stats.missing_referrals_received_from if member_stats else [])
                    if member_id in member_lookup
                ],
                'priority_connections': [
                    {
                        'id': member_id,
                        'name': member_lookup.get(member_id, 'Unknown')
                    }
                    for member_id in (member_stats.priority_connections if member_stats else [])
                    if member_id in member_lookup
                ]
            },
            'monthly_report': {
                'id': monthly_report.id,
                'month_year': monthly_report.month_year,
                'processed_at': monthly_report.processed_at
            }
        }

        return Response(result)

    except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist, Member.DoesNotExist):
        return Response(
            {'error': 'Chapter, monthly report, or member not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Member detail retrieval failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_tyfcb_data(request, chapter_id, report_id):
    """Get TYFCB data for a specific monthly report."""
    try:
        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

        tyfcb_data = {
            'inside': monthly_report.tyfcb_inside_data or {'total_amount': 0, 'count': 0, 'by_member': {}},
            'outside': monthly_report.tyfcb_outside_data or {'total_amount': 0, 'count': 0, 'by_member': {}},
            'month_year': monthly_report.month_year,
            'processed_at': monthly_report.processed_at
        }

        return Response(tyfcb_data)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except MonthlyReport.DoesNotExist:
        return Response(
            {'error': 'Monthly report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to get TYFCB data: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_member_analytics(request, chapter_id, member_name):
    """Get member analytics and recommendations based on real data."""
    try:
        from django.db.models import Sum
        from urllib.parse import unquote

        # Decode URL-encoded member name
        member_name = unquote(member_name)

        chapter = Chapter.objects.get(id=chapter_id)
        # Find member by matching their full_name property since we don't have a full_name field
        all_chapter_members = Member.objects.filter(chapter=chapter, is_active=True)
        member = None
        for m in all_chapter_members:
            if m.full_name == member_name:
                member = m
                break

        if not member:
            raise Member.DoesNotExist()

        # Get latest monthly report for this chapter
        latest_report = MonthlyReport.objects.filter(chapter=chapter).order_by('-month_year').first()

        # Get all chapter members for gap analysis
        all_members = Member.objects.filter(chapter=chapter, is_active=True).exclude(id=member.id)
        member_lookup = {m.id: m.full_name for m in all_members}

        # Get member stats from latest report if available
        member_stats = None
        if latest_report:
            try:
                member_stats = MemberMonthlyStats.objects.get(
                    member=member,
                    monthly_report=latest_report
                )
            except MemberMonthlyStats.DoesNotExist:
                pass

        # Get real analytics data from the analytics models
        referrals_given = Referral.objects.filter(giver=member).count()
        referrals_received = Referral.objects.filter(receiver=member).count()
        one_to_ones = OneToOne.objects.filter(
            models.Q(member1=member) | models.Q(member2=member)
        ).count()
        tyfcbs = TYFCB.objects.filter(receiver=member)
        total_tyfcb = tyfcbs.aggregate(total=Sum('amount'))['total'] or 0.0

        # Calculate performance score
        max_possible_otos = all_members.count()
        max_referrals = max(20, referrals_given + 10)  # Dynamic scale

        oto_score = min((one_to_ones / max_possible_otos) * 30, 30) if max_possible_otos > 0 else 0
        referral_given_score = min((referrals_given / max_referrals) * 35, 35)
        referral_received_score = min((referrals_received / max_referrals) * 35, 35)
        performance_score = round(oto_score + referral_given_score + referral_received_score)

        # Generate gap analysis - who hasn't had interactions with
        # Get all one-to-one participants for this member
        oto_partners = set()
        member_otos = OneToOne.objects.filter(
            models.Q(member1=member) | models.Q(member2=member)
        )
        for oto in member_otos:
            if oto.member1 == member:
                oto_partners.add(oto.member2.id)
            else:
                oto_partners.add(oto.member1.id)

        # Get referral partners
        referral_given_to = set(Referral.objects.filter(giver=member).values_list('receiver_id', flat=True))
        referral_received_from = set(Referral.objects.filter(receiver=member).values_list('giver_id', flat=True))

        # Calculate missing interactions
        all_member_ids = set(all_members.values_list('id', flat=True))
        missing_otos = all_member_ids - oto_partners
        missing_referrals_to = all_member_ids - referral_given_to
        missing_referrals_from = all_member_ids - referral_received_from

        # Priority connections (members with no interactions at all)
        priority_connections = missing_otos.intersection(missing_referrals_to).intersection(missing_referrals_from)

        # Generate recommendations
        recommendations = []
        if missing_otos:
            oto_names = [member_lookup[mid] for mid in list(missing_otos)[:3] if mid in member_lookup]
            if oto_names:
                recommendations.append(f"Schedule one-to-ones with {', '.join(oto_names)}")

        if missing_referrals_to:
            ref_names = [member_lookup[mid] for mid in list(missing_referrals_to)[:3] if mid in member_lookup]
            if ref_names:
                recommendations.append(f"Focus on giving referrals to {', '.join(ref_names)}")

        if missing_referrals_from:
            source_names = [member_lookup[mid] for mid in list(missing_referrals_from)[:2] if mid in member_lookup]
            if source_names:
                recommendations.append(f"Build stronger relationships with {', '.join(source_names)} to receive more referrals")

        if performance_score < 70:
            recommendations.append("Increase chapter event attendance to boost visibility")

        recommendations.append("Follow up on previous referrals to track success and build stronger connections")

        result = {
            'member': {
                'id': member.id,
                'full_name': member.full_name,
                'business_name': member.business_name,
                'classification': member.classification,
                'email': member.email,
                'phone': member.phone
            },
            'chapter': {
                'id': chapter.id,
                'name': chapter.name,
                'total_members': all_members.count() + 1  # +1 for the member themselves
            },
            'performance': {
                'referrals_given': referrals_given,
                'referrals_received': referrals_received,
                'one_to_ones': one_to_ones,
                'tyfcb_amount': float(total_tyfcb),
                'performance_score': performance_score
            },
            'gaps': {
                'missing_one_to_ones': [
                    {'id': mid, 'name': member_lookup[mid]}
                    for mid in missing_otos if mid in member_lookup
                ],
                'missing_referrals_to': [
                    {'id': mid, 'name': member_lookup[mid]}
                    for mid in missing_referrals_to if mid in member_lookup
                ],
                'missing_referrals_from': [
                    {'id': mid, 'name': member_lookup[mid]}
                    for mid in missing_referrals_from if mid in member_lookup
                ],
                'priority_connections': [
                    {'id': mid, 'name': member_lookup[mid]}
                    for mid in priority_connections if mid in member_lookup
                ]
            },
            'recommendations': recommendations,
            'completion_rates': {
                'oto_completion': round((one_to_ones / max_possible_otos) * 100) if max_possible_otos > 0 else 0,
                'referral_given_coverage': round((referrals_given / all_members.count()) * 100) if all_members.count() > 0 else 0,
                'referral_received_coverage': round((referrals_received / all_members.count()) * 100) if all_members.count() > 0 else 0
            },
            'latest_report': {
                'id': latest_report.id if latest_report else None,
                'month_year': latest_report.month_year if latest_report else None,
                'processed_at': latest_report.processed_at if latest_report else None
            }
        }

        return Response(result)

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Member.DoesNotExist:
        return Response(
            {'error': 'Member not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Member analytics failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def download_all_matrices(request, chapter_id, report_id):
    """Generate Excel file with all matrices as separate sheets."""
    try:
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import json

        chapter = Chapter.objects.get(id=chapter_id)
        monthly_report = MonthlyReport.objects.get(id=report_id, chapter=chapter)

        # Create workbook
        wb = Workbook()

        # Define styles to match old repository format
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        zero_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        bold_font = Font(bold=True)

        # Helper function to create a matrix sheet
        def create_matrix_sheet(ws, title, matrix_data, matrix_type="referral", value_formatter=None):
            ws.title = title

            if not matrix_data:
                ws['A1'] = 'No data available'
                return

            # Handle the actual matrix data format
            if 'members' in matrix_data and 'matrix' in matrix_data:
                members = matrix_data['members']
                matrix = matrix_data['matrix']
            else:
                ws['A1'] = 'Invalid matrix data format'
                return

            # Create header row (match old repository format)
            ws['A1'] = 'Giver \\ Receiver'
            ws['A1'].fill = header_fill
            ws['A1'].font = header_font
            ws['A1'].alignment = center_align
            ws['A1'].border = thin_border

            # Add column headers
            for col_idx, member_name in enumerate(members, start=2):
                cell = ws.cell(row=1, column=col_idx, value=member_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # Add summary column headers based on matrix type (match old repository format)
            if matrix_type == "combination":
                # Combination matrix has 4 summary columns
                col1 = len(members) + 2
                col2 = len(members) + 3
                col3 = len(members) + 4
                col4 = len(members) + 5

                headers = ['Neither:', 'OTO only:', 'Referral only:', 'OTO and Referral:']
                summary_cols = [col1, col2, col3, col4]

                for col, header in zip(summary_cols, headers):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
            elif matrix_type == "oto":
                # OTO matrix has 2 summary columns
                total_oto_col = len(members) + 2
                unique_oto_col = len(members) + 3

                cell = ws.cell(row=1, column=total_oto_col, value='Total OTO:')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

                cell = ws.cell(row=1, column=unique_oto_col, value=f'Unique OTO: (Total Members = {len(members)})')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            else:
                # Referral matrix has 2 summary columns
                total_given_col = len(members) + 2
                unique_given_col = len(members) + 3

                cell = ws.cell(row=1, column=total_given_col, value='Total Referrals Given:')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

                cell = ws.cell(row=1, column=unique_given_col, value='Unique Referrals Given:')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # Add data rows
            for row_idx, row_member in enumerate(members, start=2):
                # Row header
                cell = ws.cell(row=row_idx, column=1, value=row_member)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

                # Data cells with yellow highlighting for zeros
                row_total = 0
                unique_count = 0
                row_data = matrix[row_idx - 2] if row_idx - 2 < len(matrix) else []

                for col_idx, col_member in enumerate(members, start=2):
                    value = row_data[col_idx - 2] if col_idx - 2 < len(row_data) else 0

                    # Handle combination matrix format (e.g., "1/2" becomes "1/2", 0 becomes "")
                    if value_formatter:
                        display_value = value_formatter(value)
                    elif isinstance(value, str):
                        display_value = value if value != "0/0" else ''
                    else:
                        display_value = value if value else ''

                    cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
                    cell.alignment = center_align
                    cell.border = thin_border

                    # Yellow highlight for zero values (match old repository format)
                    if value == 0 and row_idx != col_idx:  # Don't highlight diagonal
                        cell.fill = zero_fill

                    # Calculate row total and unique count (for numeric values only)
                    if value and isinstance(value, (int, float)):
                        row_total += value
                        if value > 0:
                            unique_count += 1

                # Row summary values based on matrix type
                if matrix_type == "combination":
                    # For combination matrix, use simple numeric mapping
                    # - (empty/0) = Neither, 1 = OTO only, 2 = Referral only, 3 = Both
                    neither_count = 0
                    oto_only_count = 0
                    referral_only_count = 0
                    both_count = 0

                    for col_idx_inner in range(len(row_data)):
                        if row_idx - 2 != col_idx_inner:  # Don't count self-relationships
                            value = row_data[col_idx_inner] if col_idx_inner < len(row_data) else None

                            # Convert value to integer for comparison
                            if isinstance(value, str):
                                if value == '-' or value == '' or value == '0':
                                    int_value = 0
                                else:
                                    try:
                                        int_value = int(value)
                                    except (ValueError, TypeError):
                                        int_value = 0
                            elif isinstance(value, (int, float)):
                                int_value = int(value)
                            else:
                                int_value = 0

                            # Count based on simple numeric mapping
                            if int_value == 0:
                                neither_count += 1
                            elif int_value == 1:
                                oto_only_count += 1
                            elif int_value == 2:
                                referral_only_count += 1
                            elif int_value == 3:
                                both_count += 1
                            else:
                                # Any other value counts as neither
                                neither_count += 1

                    summary_values = [neither_count, oto_only_count, referral_only_count, both_count]
                    for col_offset, value in enumerate(summary_values):
                        cell = ws.cell(row=row_idx, column=len(members) + 2 + col_offset, value=value if value > 0 else '')
                        cell.alignment = center_align
                        cell.border = thin_border
                        cell.font = bold_font
                else:
                    # For referral and OTO matrices, use the calculated totals
                    cell = ws.cell(row=row_idx, column=len(members) + 2, value=row_total if row_total > 0 else '')
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.font = bold_font

                    cell = ws.cell(row=row_idx, column=len(members) + 3, value=unique_count if unique_count > 0 else '')
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.font = bold_font

            # Add summary rows based on matrix type (match old repository format)
            if matrix_type == "combination":
                # Combination matrix doesn't have summary rows in the old format
                pass
            else:
                total_received_row = len(members) + 2
                unique_received_row = len(members) + 3

                # Summary row labels based on matrix type
                if matrix_type == "oto":
                    total_label = 'Total OTO Received:'
                    unique_label = 'Unique OTO Received:'
                else:
                    total_label = 'Total Referrals Received:'
                    unique_label = 'Unique Referrals Received:'

                # Total received row
                cell = ws.cell(row=total_received_row, column=1, value=total_label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

                # Unique received row
                cell = ws.cell(row=unique_received_row, column=1, value=unique_label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

                # Calculate column totals and unique counts for non-combination matrices
                grand_total = 0
                grand_unique = 0

                for col_idx in range(len(members)):
                    col_total = 0
                    col_unique = 0

                    for row_idx in range(len(matrix)):
                        if row_idx < len(matrix) and col_idx < len(matrix[row_idx]):
                            value = matrix[row_idx][col_idx]
                            if value and isinstance(value, (int, float)):
                                col_total += value
                                grand_total += value
                                if value > 0:
                                    col_unique += 1
                                    grand_unique += 1

                    # Total received for this member
                    cell = ws.cell(row=total_received_row, column=col_idx + 2, value=col_total if col_total else '')
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.font = bold_font

                    # Unique received for this member
                    cell = ws.cell(row=unique_received_row, column=col_idx + 2, value=col_unique if col_unique else '')
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.font = bold_font

                # Grand totals in summary columns
                summary_col1 = len(members) + 2
                summary_col2 = len(members) + 3

                cell = ws.cell(row=total_received_row, column=summary_col1, value=grand_total if grand_total else '')
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = bold_font

                cell = ws.cell(row=unique_received_row, column=summary_col2, value=grand_unique if grand_unique else '')
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = bold_font

            # Adjust column widths (match old repository format)
            num_summary_cols = 4 if matrix_type == "combination" else 2
            for col in range(1, len(members) + 2 + num_summary_cols):
                ws.column_dimensions[get_column_letter(col)].width = 15
            ws.column_dimensions['A'].width = 25  # Wider for names
            # Make summary columns wider for longer headers
            for i in range(num_summary_cols):
                col_letter = get_column_letter(len(members) + 2 + i)
                ws.column_dimensions[col_letter].width = 20

        # Remove default sheet
        wb.remove(wb.active)

        # Create Referral Matrix sheet
        if monthly_report.referral_matrix_data:
            ws_referral = wb.create_sheet("Referral Matrix")
            create_matrix_sheet(ws_referral, "Referral Matrix", monthly_report.referral_matrix_data, "referral")

        # Create One-to-One Matrix sheet
        if monthly_report.oto_matrix_data:
            ws_oto = wb.create_sheet("One-to-One Matrix")
            create_matrix_sheet(ws_oto, "One-to-One Matrix", monthly_report.oto_matrix_data, "oto")

        # Create Combination Matrix sheet
        if monthly_report.combination_matrix_data:
            ws_combo = wb.create_sheet("Combination Matrix")
            # For combination matrix, format the values differently
            def format_combo(value):
                if not value or value == '0/0':
                    return ''
                return value
            create_matrix_sheet(ws_combo, "Combination Matrix", monthly_report.combination_matrix_data, "combination", format_combo)

        # If no matrices available, add a notice sheet
        if not wb.worksheets:
            ws = wb.create_sheet("No Data")
            ws['A1'] = "No matrix data available for this report"

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{chapter.name.replace(' ', '_')}_Matrices_{monthly_report.month_year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except MonthlyReport.DoesNotExist:
        return Response(
            {'error': 'Monthly report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to generate Excel: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def create_chapter(request):
    """Create a new chapter using ChapterService."""
    try:
        # Use ChapterService for consistent chapter creation
        chapter, created = ChapterService.get_or_create_chapter(
            name=request.data.get('name', ''),
            location=request.data.get('location', 'Dubai'),
            meeting_day=request.data.get('meeting_day', ''),
            meeting_time=request.data.get('meeting_time'),
        )

        return Response(
            ChapterSerializer(chapter).data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )

    except ValidationError as e:
        return Response(
            {'error': 'Invalid data', 'details': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to create chapter: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([AllowAny])
def delete_chapter(request, chapter_id):
    """Delete a chapter using ChapterService."""
    try:
        result = ChapterService.delete_chapter(chapter_id)
        return Response(
            {'message': f'Chapter "{result["chapter_name"]}" deleted successfully',
             'members_deleted': result['members_deleted']},
            status=status.HTTP_200_OK
        )
    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to delete chapter: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def create_member(request, chapter_id):
    """Create a new member using MemberService."""
    try:
        chapter = ChapterService.get_chapter(chapter_id)
    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    try:
        # Use MemberService for consistent member creation
        member, created = MemberService.get_or_create_member(
            chapter=chapter,
            first_name=request.data.get('first_name', ''),
            last_name=request.data.get('last_name', ''),
            business_name=request.data.get('business_name', ''),
            classification=request.data.get('classification', ''),
            email=request.data.get('email', ''),
            phone=request.data.get('phone', ''),
            is_active=request.data.get('is_active', True),
            joined_date=request.data.get('joined_date'),
        )

        return Response(
            MemberSerializer(member).data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )

    except ValidationError as e:
        return Response(
            {'error': 'Invalid data', 'details': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to create member: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['PUT', 'PATCH'])
@permission_classes([AllowAny])
def update_member(request, chapter_id, member_id):
    """Update a member using MemberService."""
    try:
        chapter = ChapterService.get_chapter(chapter_id)
        member = MemberService.get_member(member_id)

        # Verify member belongs to chapter
        if member.chapter.id != chapter.id:
            return Response(
                {'error': 'Member not found in this chapter'},
                status=status.HTTP_404_NOT_FOUND
            )

    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Member.DoesNotExist:
        return Response(
            {'error': 'Member not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    try:
        # Use MemberService for consistent member updates
        updated_member, was_updated = MemberService.update_member(
            member_id,
            **request.data
        )

        return Response(
            MemberSerializer(updated_member).data,
            status=status.HTTP_200_OK
        )

    except ValidationError as e:
        return Response(
            {'error': 'Invalid data', 'details': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to update member: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([AllowAny])
def delete_member(request, chapter_id, member_id):
    """Delete a member using MemberService."""
    try:
        chapter = ChapterService.get_chapter(chapter_id)
        member = MemberService.get_member(member_id)

        # Verify member belongs to chapter
        if member.chapter.id != chapter.id:
            return Response(
                {'error': 'Member not found in this chapter'},
                status=status.HTTP_404_NOT_FOUND
            )

        result = MemberService.delete_member(member_id)
        return Response(
            {'message': f'Member "{result["member_name"]}" deleted successfully',
             'referrals_deleted': result['referrals_deleted'],
             'one_to_ones_deleted': result['one_to_ones_deleted'],
             'tyfcbs_deleted': result['tyfcbs_deleted']},
            status=status.HTTP_200_OK
        )
    except Chapter.DoesNotExist:
        return Response(
            {'error': 'Chapter not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Member.DoesNotExist:
        return Response(
            {'error': 'Member not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to delete member: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Matrix Comparison Endpoints

@api_view(['GET'])
@permission_classes([AllowAny])
def compare_referral_matrices(request, chapter_id, report_id, previous_report_id):
    """Compare referral matrices between two monthly reports."""
    try:
        # Get both reports
        current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
        previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

        # Perform comparison
        comparison = ComparisonService.compare_referral_matrices(
            current_report.referral_matrix_data,
            previous_report.referral_matrix_data
        )

        return Response({
            'current_report': {
                'id': current_report.id,
                'month_year': current_report.month_year
            },
            'previous_report': {
                'id': previous_report.id,
                'month_year': previous_report.month_year
            },
            'comparison': comparison
        }, status=status.HTTP_200_OK)

    except MonthlyReport.DoesNotExist:
        return Response(
            {'error': 'One or both reports not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to compare referral matrices: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def compare_oto_matrices(request, chapter_id, report_id, previous_report_id):
    """Compare one-to-one matrices between two monthly reports."""
    try:
        # Get both reports
        current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
        previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

        # Perform comparison
        comparison = ComparisonService.compare_oto_matrices(
            current_report.oto_matrix_data,
            previous_report.oto_matrix_data
        )

        return Response({
            'current_report': {
                'id': current_report.id,
                'month_year': current_report.month_year
            },
            'previous_report': {
                'id': previous_report.id,
                'month_year': previous_report.month_year
            },
            'comparison': comparison
        }, status=status.HTTP_200_OK)

    except MonthlyReport.DoesNotExist:
        return Response(
            {'error': 'One or both reports not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to compare one-to-one matrices: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def compare_combination_matrices(request, chapter_id, report_id, previous_report_id):
    """Compare combination matrices between two monthly reports."""
    try:
        # Get both reports
        current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
        previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

        # Perform comparison
        comparison = ComparisonService.compare_combination_matrices(
            current_report.combination_matrix_data,
            previous_report.combination_matrix_data
        )

        return Response({
            'current_report': {
                'id': current_report.id,
                'month_year': current_report.month_year
            },
            'previous_report': {
                'id': previous_report.id,
                'month_year': previous_report.month_year
            },
            'comparison': comparison
        }, status=status.HTTP_200_OK)

    except MonthlyReport.DoesNotExist:
        return Response(
            {'error': 'One or both reports not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to compare combination matrices: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def compare_reports(request, chapter_id, report_id, previous_report_id):
    """Get comprehensive comparison between two monthly reports including all matrices and insights."""
    try:
        # Get both reports
        current_report = MonthlyReport.objects.get(id=report_id, chapter_id=chapter_id)
        previous_report = MonthlyReport.objects.get(id=previous_report_id, chapter_id=chapter_id)

        # Perform comprehensive comparison
        comparison_data = ComparisonService.compare_monthly_reports(
            current_report,
            previous_report
        )

        return Response(comparison_data, status=status.HTTP_200_OK)

    except MonthlyReport.DoesNotExist:
        return Response(
            {'error': 'One or both reports not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except ValueError as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {'error': f'Failed to compare reports: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
