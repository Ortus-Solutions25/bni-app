"""
MonthlyReport ViewSet - RESTful API for Monthly Report management
"""
from django.http import HttpResponse
from django.db.models import Count, Sum
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from chapters.models import Chapter
from members.models import Member
from reports.models import MonthlyReport, MemberMonthlyStats
from analytics.models import TYFCB


class MonthlyReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for MonthlyReport operations.

    Provides:
    - list: Get all monthly reports for a chapter
    - destroy: Delete a monthly report
    - member_detail: Get detailed member information with missing interaction lists
    - tyfcb_data: Get TYFCB data for a specific report
    """
    queryset = MonthlyReport.objects.all()
    permission_classes = [AllowAny]  # TODO: Add proper authentication

    def list(self, request, chapter_id=None):
        """
        Get all monthly reports for a specific chapter.

        Returns list of monthly reports with metadata including:
        - Report IDs and dates
        - File information
        - Matrix availability flags
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_reports = MonthlyReport.objects.filter(chapter=chapter).order_by('-month_year')

            result = []
            for report in monthly_reports:
                try:
                    # Safely access file fields (they are strings, not actual files)
                    slip_file = str(report.slip_audit_file) if report.slip_audit_file else None
                    member_file = str(report.member_names_file) if report.member_names_file else None

                    result.append({
                        'id': report.id,
                        'month_year': report.month_year,
                        'uploaded_at': report.uploaded_at.isoformat() if report.uploaded_at else None,
                        'processed_at': report.processed_at.isoformat() if report.processed_at else None,
                        'slip_audit_file': slip_file,
                        'member_names_file': member_file,
                        'has_referral_matrix': bool(report.referral_matrix_data),
                        'has_oto_matrix': bool(report.oto_matrix_data),
                        'has_combination_matrix': bool(report.combination_matrix_data)
                    })
                except Exception as e:
                    # Log the error but continue processing other reports
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Error processing report {report.id}: {str(e)}")
                    continue

            return Response(result)

        except Chapter.DoesNotExist:
            return Response(
                {'error': 'Chapter not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Monthly reports retrieval failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def destroy(self, request, pk=None, chapter_id=None):
        """
        Delete a monthly report.

        Also deletes associated files from storage.
        """
        # Skip authentication check for now (TODO: Add proper auth later)
        # if request.user and not request.user.is_authenticated:
        #     return Response(
        #         {'error': 'Authentication required'},
        #         status=status.HTTP_401_UNAUTHORIZED
        #     )

        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Delete the report (files are just filenames stored as strings, no actual files to delete)
            monthly_report.delete()

            return Response({'message': 'Monthly report deleted successfully'})

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
            return Response(
                {'error': 'Chapter or monthly report not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Failed to delete report: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='members/(?P<member_id>[^/.]+)')
    def member_detail(self, request, pk=None, chapter_id=None, member_id=None):
        """
        Get detailed member information including missing interaction lists.

        Returns member stats and lists of:
        - Missing one-to-ones
        - Missing referrals (given and received)
        - Priority connections (members appearing in multiple missing lists)
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)
            member = Member.objects.get(id=member_id, chapter=chapter)

            try:
                member_stats = MemberMonthlyStats.objects.get(
                    member=member,
                    monthly_report=monthly_report
                )
            except MemberMonthlyStats.DoesNotExist:
                # If no stats exist, return basic member info with empty lists
                member_stats = None

            # Get all chapter members for name resolution
            chapter_members = Member.objects.filter(chapter=chapter, is_active=True)
            member_lookup = {m.id: m.full_name for m in chapter_members}

            result = {
                'member': {
                    'id': member.id,
                    'full_name': member.full_name,
                    'first_name': member.first_name,
                    'last_name': member.last_name,
                    'business_name': member.business_name,
                    'classification': member.classification,
                    'email': member.email,
                    'phone': member.phone
                },
                'stats': {
                    'referrals_given': member_stats.referrals_given if member_stats else 0,
                    'referrals_received': member_stats.referrals_received if member_stats else 0,
                    'one_to_ones_completed': member_stats.one_to_ones_completed if member_stats else 0,
                    'tyfcb_inside_amount': float(member_stats.tyfcb_inside_amount) if member_stats else 0.0,
                    'tyfcb_outside_amount': float(member_stats.tyfcb_outside_amount) if member_stats else 0.0
                },
                'missing_interactions': {
                    'missing_otos': [
                        {
                            'id': member_id,
                            'name': member_lookup.get(member_id, 'Unknown')
                        }
                        for member_id in (member_stats.missing_otos if member_stats else [])
                        if member_id in member_lookup
                    ],
                    'missing_referrals_given_to': [
                        {
                            'id': member_id,
                            'name': member_lookup.get(member_id, 'Unknown')
                        }
                        for member_id in (member_stats.missing_referrals_given_to if member_stats else [])
                        if member_id in member_lookup
                    ],
                    'missing_referrals_received_from': [
                        {
                            'id': member_id,
                            'name': member_lookup.get(member_id, 'Unknown')
                        }
                        for member_id in (member_stats.missing_referrals_received_from if member_stats else [])
                        if member_id in member_lookup
                    ],
                    'priority_connections': [
                        {
                            'id': member_id,
                            'name': member_lookup.get(member_id, 'Unknown')
                        }
                        for member_id in (member_stats.priority_connections if member_stats else [])
                        if member_id in member_lookup
                    ]
                },
                'monthly_report': {
                    'id': monthly_report.id,
                    'month_year': monthly_report.month_year,
                    'processed_at': monthly_report.processed_at
                }
            }

            return Response(result)

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist, Member.DoesNotExist):
            return Response(
                {'error': 'Chapter, monthly report, or member not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Member detail retrieval failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='tyfcb')
    def tyfcb_data(self, request, pk=None, chapter_id=None):
        """
        Get TYFCB data for a specific monthly report.

        Returns inside and outside TYFCB data with totals and per-member breakdowns.
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            tyfcb_data = {
                'inside': monthly_report.tyfcb_inside_data or {'total_amount': 0, 'count': 0, 'by_member': {}},
                'outside': monthly_report.tyfcb_outside_data or {'total_amount': 0, 'count': 0, 'by_member': {}},
                'month_year': monthly_report.month_year,
                'processed_at': monthly_report.processed_at
            }

            return Response(tyfcb_data)

        except Chapter.DoesNotExist:
            return Response(
                {'error': 'Chapter not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {'error': 'Monthly report not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to get TYFCB data: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='download-matrices')
    def download_matrices(self, request, pk=None, chapter_id=None):
        """
        Generate and download Excel file with all matrices.

        Creates a workbook with separate sheets for:
        - Referral Matrix
        - One-to-One Matrix
        - Combination Matrix
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Create a new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")

            # Helper function to create matrix sheet
            def create_matrix_sheet(sheet_name, matrix_data, include_aggregates=False):
                if not matrix_data or 'members' not in matrix_data or 'matrix' not in matrix_data:
                    return None

                ws = wb.create_sheet(title=sheet_name)
                members = matrix_data['members']
                matrix = matrix_data['matrix']  # This is a 2D list

                # Determine header columns
                num_cols = len(members) + (4 if include_aggregates else 0)

                # Write header row
                ws.cell(1, 1, "From \\ To").font = header_font
                ws.cell(1, 1).fill = header_fill
                ws.cell(1, 1).alignment = center_align

                for col_idx, member in enumerate(members, start=2):
                    cell = ws.cell(1, col_idx, member)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

                # Add aggregate column headers for combination matrix
                if include_aggregates:
                    aggregate_headers = ["Neither", "OTO Only", "Referral Only", "Both"]
                    aggregate_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    for idx, header in enumerate(aggregate_headers):
                        col = len(members) + 2 + idx
                        cell = ws.cell(1, col, header)
                        cell.font = header_font
                        cell.fill = aggregate_fill
                        cell.alignment = center_align
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

                # Write data rows - matrix is a 2D list
                for row_idx, from_member in enumerate(members):
                    # Row header
                    cell = ws.cell(row_idx + 2, 1, from_member)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = center_align

                    # Matrix values from the 2D list
                    row_total_given = 0
                    if row_idx < len(matrix):
                        row_data = matrix[row_idx]
                        for col_idx, value in enumerate(row_data):
                            cell = ws.cell(row_idx + 2, col_idx + 2, value)
                            cell.alignment = center_align
                            row_total_given += value if isinstance(value, (int, float)) else 0

                            # Color coding for values
                            if value > 0:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                    # Add aggregates for combination matrix
                    if include_aggregates:
                        # Count each type based on legend:
                        # 0 = Neither, 1 = OTO Only, 2 = Referral Only, 3 = Both
                        row_data = matrix[row_idx] if row_idx < len(matrix) else []

                        neither_count = sum(1 for v in row_data if v == 0)
                        oto_only_count = sum(1 for v in row_data if v == 1)
                        ref_only_count = sum(1 for v in row_data if v == 2)
                        both_count = sum(1 for v in row_data if v == 3)

                        # Write aggregate values
                        aggregate_values = [neither_count, oto_only_count, ref_only_count, both_count]
                        for idx, value in enumerate(aggregate_values):
                            col = len(members) + 2 + idx
                            cell = ws.cell(row_idx + 2, col, value)
                            cell.alignment = center_align
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

                ws.column_dimensions['A'].width = 20
                return ws

            # Create sheets for each matrix type
            if monthly_report.referral_matrix_data:
                create_matrix_sheet("Referral Matrix", monthly_report.referral_matrix_data)

            if monthly_report.oto_matrix_data:
                create_matrix_sheet("One-to-One Matrix", monthly_report.oto_matrix_data)

            if monthly_report.combination_matrix_data:
                create_matrix_sheet("Combination Matrix", monthly_report.combination_matrix_data, include_aggregates=True)

            # Create TYFCB sheet
            if monthly_report.tyfcb_inside_data or monthly_report.tyfcb_outside_data:
                ws_tyfcb = wb.create_sheet("TYFCB Report")

                # Header
                ws_tyfcb.cell(1, 1, "TYFCB Report").font = Font(bold=True, size=14)
                ws_tyfcb.merge_cells('A1:D1')

                row = 3
                # Inside Chapter TYFCB
                if monthly_report.tyfcb_inside_data:
                    inside = monthly_report.tyfcb_inside_data
                    ws_tyfcb.cell(row, 1, "Within Chapter").font = Font(bold=True, size=12)
                    ws_tyfcb.cell(row, 1).fill = header_fill
                    row += 1
                    ws_tyfcb.cell(row, 1, f"Total Amount: AED {inside.get('total_amount', 0):,.2f}").font = Font(bold=True)
                    ws_tyfcb.cell(row, 3, f"Total TYFCBs: {inside.get('count', 0)}").font = Font(bold=True)
                    row += 2

                    # By member breakdown - use data from JSON field
                    ws_tyfcb.cell(row, 1, "Member").font = header_font
                    ws_tyfcb.cell(row, 1).fill = header_fill
                    ws_tyfcb.cell(row, 2, "Amount (AED)").font = header_font
                    ws_tyfcb.cell(row, 2).fill = header_fill
                    row += 1

                    # Get by_member data and sort by amount (descending)
                    by_member = inside.get('by_member', {})
                    sorted_members = sorted(by_member.items(), key=lambda x: x[1], reverse=True)

                    for member_name, amount in sorted_members:
                        if amount > 0:  # Only show members with TYFCB
                            ws_tyfcb.cell(row, 1, member_name)
                            ws_tyfcb.cell(row, 2, f"{float(amount):,.2f}")
                            row += 1

                    row += 2

                # Outside Chapter TYFCB
                if monthly_report.tyfcb_outside_data:
                    outside = monthly_report.tyfcb_outside_data
                    ws_tyfcb.cell(row, 1, "Outside Chapter").font = Font(bold=True, size=12)
                    ws_tyfcb.cell(row, 1).fill = header_fill
                    row += 1
                    ws_tyfcb.cell(row, 1, f"Total Amount: AED {outside.get('total_amount', 0):,.2f}").font = Font(bold=True)
                    ws_tyfcb.cell(row, 3, f"Total TYFCBs: {outside.get('count', 0)}").font = Font(bold=True)
                    row += 2

                    # By member breakdown - use data from JSON field
                    ws_tyfcb.cell(row, 1, "Member").font = header_font
                    ws_tyfcb.cell(row, 1).fill = header_fill
                    ws_tyfcb.cell(row, 2, "Amount (AED)").font = header_font
                    ws_tyfcb.cell(row, 2).fill = header_fill
                    row += 1

                    # Get by_member data and sort by amount (descending)
                    by_member = outside.get('by_member', {})
                    sorted_members = sorted(by_member.items(), key=lambda x: x[1], reverse=True)

                    for member_name, amount in sorted_members:
                        if amount > 0:  # Only show members with TYFCB
                            ws_tyfcb.cell(row, 1, member_name)
                            ws_tyfcb.cell(row, 2, f"{float(amount):,.2f}")
                            row += 1

                # Set column widths
                ws_tyfcb.column_dimensions['A'].width = 30
                ws_tyfcb.column_dimensions['B'].width = 20
                ws_tyfcb.column_dimensions['C'].width = 15
                ws_tyfcb.column_dimensions['D'].width = 15

            # If no matrices were created, add an info sheet
            if len(wb.sheetnames) == 0:
                ws = wb.create_sheet("Info")
                ws['A1'] = "No matrix data available for this report"
                ws['A1'].font = Font(bold=True)

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create HTTP response
            filename = f"{chapter.name.replace(' ', '_')}_Matrices_{monthly_report.month_year}.xlsx"
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Chapter.DoesNotExist:
            return Response(
                {'error': 'Chapter not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {'error': 'Monthly report not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to generate Excel file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
